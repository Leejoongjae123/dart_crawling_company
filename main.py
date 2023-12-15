import openpyxl
import pandas as pd
from pyautogui import size
from selenium import webdriver
from selenium.webdriver.chrome.options import Options
import subprocess
import shutil
from selenium.webdriver.common.by import By
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.common.desired_capabilities import DesiredCapabilities
from bs4 import BeautifulSoup
import time
import datetime
import pyautogui
import pyperclip
import csv
import sys
import os
import math
import requests
import re
import random
import chromedriver_autoinstaller
from PyQt5.QtWidgets import QWidget, QApplication, QTreeView, QFileSystemModel, QVBoxLayout, QPushButton, QInputDialog, \
    QLineEdit, QMainWindow, QMessageBox, QFileDialog
from PyQt5.QtCore import QCoreApplication
from selenium.webdriver import ActionChains
from datetime import datetime, date, timedelta
import numpy
import datetime
from window import Ui_MainWindow
from PyQt5.QtWidgets import *
from PyQt5.QtCore import *
import json
import pprint
from selenium.webdriver.chrome.service import Service
from webdriver_manager.chrome import ChromeDriverManager
from PyQt5.QtCore import QDate
from openpyxl.styles import Alignment
from openpyxl.styles import Font
from openpyxl.styles import PatternFill

# Function to format the numbers
def format_number(num):
    # If the second decimal is zero, display up to three decimal places
    if round(num, 2) == round(num, 1):
        return f"{num:.3f}"
    # Otherwise, display up to two decimal places
    else:
        return f"{num:.2f}"
def GetID(category,year,month,day):
    count=1
    endFlag=False
    dataList=[]
    while True:
        cookies = {
            'WMONID': 'D1vNDyvY7sc',
            'JSESSIONID': 'wf8kYTfyDq5Mul19Fkq5QOga2WVHVnPjvKKqxFUu4K6bjaEgn97eUgtpcQFBtvIv.ZG1fZGFydC9kYXJ0MV9kYXJ0X21zMQ==',
        }

        headers = {
            'Accept': 'text/html, */*; q=0.01',
            'Accept-Language': 'ko-KR,ko;q=0.9,en-US;q=0.8,en;q=0.7',
            'Connection': 'keep-alive',
            'Content-Type': 'application/x-www-form-urlencoded; charset=UTF-8',
            # 'Cookie': 'WMONID=D1vNDyvY7sc; JSESSIONID=wf8kYTfyDq5Mul19Fkq5QOga2WVHVnPjvKKqxFUu4K6bjaEgn97eUgtpcQFBtvIv.ZG1fZGFydC9kYXJ0MV9kYXJ0X21zMQ==',
            'Origin': 'https://dart.fss.or.kr',
            'Referer': 'https://dart.fss.or.kr/dsab007/main.do',
            'Sec-Fetch-Dest': 'empty',
            'Sec-Fetch-Mode': 'cors',
            'Sec-Fetch-Site': 'same-origin',
            'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/119.0.0.0 Safari/537.36',
            'X-Requested-With': 'XMLHttpRequest',
            'sec-ch-ua': '"Google Chrome";v="119", "Chromium";v="119", "Not?A_Brand";v="24"',
            'sec-ch-ua-mobile': '?0',
            'sec-ch-ua-platform': '"Windows"',
        }

        timeNow=datetime.datetime.now().strftime("%Y%m%d")
        timePast=(datetime.datetime.now()-datetime.timedelta(days=30)).strftime("%Y%m%d")
        if category=="임원":
            searchCategory='임원ㆍ주요주주특정증권등소유상황보고서'
        else:
            searchCategory = '최대주주등소유주식변동신고서'
        data = {
            'currentPage': count,
            'maxResults': '15',
            'maxLinks': '10',
            'sort': '',
            'series': '',
            'textCrpCik': '',
            'lateKeyword': '',
            'keyword': '',
            'reportNamePopYn': '',
            'textkeyword': '',
            'businessCode': 'all',
            'autoSearch': 'N',
            'option': 'report',
            'textCrpNm': '',
            'reportName': searchCategory,
            'tocSrch': '',
            'textCrpNm2': '',
            'textPresenterNm': '',
            'startDate': timePast,
            'endDate': timeNow,
            'decadeType': '',
            'finalReport': 'recent',
            'businessNm': '전체',
            'corporationType': 'all',
            'closingAccountsMonth': 'all',
            'tocSrch2': '',
        }


        response = requests.post('https://dart.fss.or.kr/dsab007/detailSearch.ax', cookies=cookies, headers=headers, data=data)
        # print(response.text)
        soup=BeautifulSoup(response.text,'lxml')
        # print(soup.prettify())
        table=soup.find("table",attrs={'class':'tbList'})
        trs=table.find_all('tr')
        for index,tr in enumerate(trs):
            if index==0:
                continue
            try:
                title=tr.find_all("td")[1].find("a").get_text().replace("\n","").strip()
            except:
                title=""
            print("title:",title)

            try:
                reportId=tr.find_all("td")[2].find("a")['href'].replace("\n","").split("?")[-1].replace("rcpNo=","").strip()
            except:
                reportId=""
            print("reportId:", reportId)
            try:
                regiDate=tr.find_all("td")[-2].get_text()
            except:
                regiDate=""
            print("regiDate:",regiDate)
            regiTimestamp=datetime.datetime.strptime(regiDate,'%Y.%m.%d').timestamp()

            timeStart=datetime.datetime(year,month,day).timestamp()
            if regiTimestamp<timeStart:
                endFlag=True
                break

            data={'title':title,'reportId':reportId,'regiDate':regiDate}
            dataList.append(data)
            print("=====================")
        if endFlag==True:
            break
        count+=1
        time.sleep(random.randint(5,10)*0.1)
    with open('idList.json', 'w',encoding='utf-8-sig') as f:
        json.dump(dataList, f, indent=2,ensure_ascii=False)
    return dataList
def GetPreDetali(idInfo):
    cookies = {
        'WMONID': 'D1vNDyvY7sc',
        'JSESSIONID': 'XChPHxuScxLNn7YBSj3Gw2NtPdBSQbAxlu8FdkAvSYKynqfanbzuMiu2YqPZyF0x.ZG1fZGFydC9kYXJ0MV9kYXJ0X21zMQ==',
    }

    headers = {
        'Accept': 'text/html,application/xhtml+xml,application/xml;q=0.9,image/avif,image/webp,image/apng,*/*;q=0.8,application/signed-exchange;v=b3;q=0.7',
        'Accept-Language': 'ko-KR,ko;q=0.9,en-US;q=0.8,en;q=0.7',
        'Cache-Control': 'max-age=0',
        'Connection': 'keep-alive',
        # 'Cookie': 'WMONID=D1vNDyvY7sc; JSESSIONID=XChPHxuScxLNn7YBSj3Gw2NtPdBSQbAxlu8FdkAvSYKynqfanbzuMiu2YqPZyF0x.ZG1fZGFydC9kYXJ0MV9kYXJ0X21zMQ==',
        'Referer': 'https://dart.fss.or.kr/dsab007/main.do',
        'Sec-Fetch-Dest': 'document',
        'Sec-Fetch-Mode': 'navigate',
        'Sec-Fetch-Site': 'same-origin',
        'Sec-Fetch-User': '?1',
        'Upgrade-Insecure-Requests': '1',
        'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/119.0.0.0 Safari/537.36',
        'sec-ch-ua': '"Google Chrome";v="119", "Chromium";v="119", "Not?A_Brand";v="24"',
        'sec-ch-ua-mobile': '?0',
        'sec-ch-ua-platform': '"Windows"',
    }

    params = {
        'rcpNo': idInfo['reportId'],
    }

    response = requests.get('https://dart.fss.or.kr/dsaf001/main.do', params=params, cookies=cookies, headers=headers)
    soup=BeautifulSoup(response.text,'lxml')

    #데이타를 미리 준비

    data1={'rcpNo':idInfo['reportId']}
    data2={'rcpNo':idInfo['reportId']}
    data3={'rcpNo':idInfo['reportId']}
    data4={'rcpNo':idInfo['reportId']}
    data5={'rcpNo':idInfo['reportId']}

    # 정규식 패턴을 정의합니다.
    pattern = r"node1\['text'\] = \"([^']+)\""

    # 정규식 패턴과 매치되는 부분을 찾습니다.
    matches = re.findall(pattern, response.text)

    # 매치된 결과를 출력합니다.
    for index,match in enumerate(matches):
        # print("node1['text'] =", match)
        if index==0:
            data1['text']=match
        if index==1:
            data2['text'] = match
        if index==2:
            data3['text'] = match
        if index==3:
            data4['text'] = match
        try:
            if index==4:
                data5['text'] = match
        except:
            print("4번째없음")

    # 정규식 패턴을 정의합니다.
    pattern = r"node1\['dcmNo'\] = \"([^']+)\""

    # 정규식 패턴과 매치되는 부분을 찾습니다.
    matches = re.findall(pattern, response.text)

    # 매치된 결과를 출력합니다.
    for index,match in enumerate(matches):
        # print("node1['text'] =", match)
        if index==0:
            data1['dcmNo']=match.replace("node1['dcmNo'] =","")
        if index==1:
            data2['dcmNo'] = match.replace("node1['dcmNo'] =", "")
        if index==2:
            data3['dcmNo'] = match.replace("node1['dcmNo'] =", "")
        if index==3:
            data4['dcmNo'] = match.replace("node1['dcmNo'] =", "")
        try:
            if index==4:
                data5['dcmNo'] = match.replace("node1['dcmNo'] =", "")
        except:
            print("5번째없음")

    # 정규식 패턴을 정의합니다.
    pattern = r"node1\['eleId'\] = \"([^']+)\""

    # 정규식 패턴과 매치되는 부분을 찾습니다.
    matches = re.findall(pattern, response.text)

    # 매치된 결과를 출력합니다.
    for index,match in enumerate(matches):
        # print("node1['text'] =", match)
        if index==0:
            data1['eleId']=match
        if index==1:
            data2['eleId'] = match
        if index==2:
            data3['eleId'] = match
        if index==3:
            data4['eleId'] = match
        try:
            if index==4:
                data5['eleId'] = match
        except:
            print("5번째없음")

    # 정규식 패턴을 정의합니다.
    pattern = r"node1\['offset'\] = \"([^']+)\""

    # 정규식 패턴과 매치되는 부분을 찾습니다.
    matches = re.findall(pattern, response.text)

    # 매치된 결과를 출력합니다.
    for index,match in enumerate(matches):
        # print("node1['text'] =", match)
        if index==0:
            data1['offset']=match
        if index==1:
            data2['offset'] = match
        if index==2:
            data3['offset'] = match
        if index==3:
            data4['offset'] = match
        try:
            if index==4:
                data5['offset'] = match
        except:
            print("5번째없음")

    # 정규식 패턴을 정의합니다.
    pattern = r"node1\['length'\] = \"([^']+)\""

    # 정규식 패턴과 매치되는 부분을 찾습니다.
    matches = re.findall(pattern, response.text)

    # 매치된 결과를 출력합니다.
    for index,match in enumerate(matches):
        # print("node1['text'] =", match)
        if index==0:
            data1['length']=match
        if index==1:
            data2['length'] = match
        if index==2:
            data3['length'] = match
        if index==3:
            data4['length'] = match
        try:
            if index==4:
                data5['length'] = match
        except:
            print("5번째없음")

    # pprint.pprint(data1)
    # pprint.pprint(data2)
    # pprint.pprint(data3)
    # pprint.pprint(data4)

    dataList=[]
    # dataList.append(data1)
    dataList.append(data2)
    dataList.append(data3)
    dataList.append(data4)
    try:
        if data5['text']:
            dataList.append(data5)
    except:
        print("5번째없음")
    print("dataList:",dataList,"/ dataList_TYPE:",type(dataList))
    pprint.pprint(dataList)
    return dataList
def GetInfos(data):
    cookies = {
        'WMONID': 'D1vNDyvY7sc',
        'JSESSIONID': 'XChPHxuScxLNn7YBSj3Gw2NtPdBSQbAxlu8FdkAvSYKynqfanbzuMiu2YqPZyF0x.ZG1fZGFydC9kYXJ0MV9kYXJ0X21zMQ==',
    }

    headers = {
        'Accept': 'text/html,application/xhtml+xml,application/xml;q=0.9,image/avif,image/webp,image/apng,*/*;q=0.8,application/signed-exchange;v=b3;q=0.7',
        'Accept-Language': 'ko-KR,ko;q=0.9,en-US;q=0.8,en;q=0.7',
        'Connection': 'keep-alive',
        # 'Cookie': 'WMONID=D1vNDyvY7sc; JSESSIONID=XChPHxuScxLNn7YBSj3Gw2NtPdBSQbAxlu8FdkAvSYKynqfanbzuMiu2YqPZyF0x.ZG1fZGFydC9kYXJ0MV9kYXJ0X21zMQ==',
        'Referer': 'https://dart.fss.or.kr/dsaf001/main.do?rcpNo=20231130000698',
        'Sec-Fetch-Dest': 'iframe',
        'Sec-Fetch-Mode': 'navigate',
        'Sec-Fetch-Site': 'same-origin',
        'Sec-Fetch-User': '?1',
        'Upgrade-Insecure-Requests': '1',
        'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/119.0.0.0 Safari/537.36',
        'sec-ch-ua': '"Google Chrome";v="119", "Chromium";v="119", "Not?A_Brand";v="24"',
        'sec-ch-ua-mobile': '?0',
        'sec-ch-ua-platform': '"Windows"',
    }

    params = {
        'rcpNo': data['rcpNo'],
        'dcmNo': data['dcmNo'],
        'eleId': data['eleId'],
        'offset': data['offset'],
        'length': data['length'],
        'dtd': 'dart3.xsd',
    }
    print("params:",params,"/ params_TYPE:",type(params))

    response = requests.get('https://dart.fss.or.kr/report/viewer.do', params=params, cookies=cookies, headers=headers)
    # print(response.text)
    soup=BeautifulSoup(response.text,'lxml')
    company=""
    name=""
    chair=""
    saleType = ''
    totalValue=0
    regiDateList=[]
    url='https://dart.fss.or.kr/dsaf001/main.do?rcpNo={}'.format(data['rcpNo'])
    if data['text'].find("발행회사에")>=0:
        try:
            company=soup.find_all('td')[1].get_text()
        except:
            company=""
        print("company:",company,"/ company_TYPE:",type(company))
        data={'company':company,'url':url}
        return data
    elif data['text'].find("보고자에 관한")>=0:
        tds=soup.find_all("td")
        for index,td in enumerate(tds):
            if td.get_text().find("성명")>=0:
                name=tds[index+2].get_text()
                print("name:",name,"/ name_TYPE:",type(name))
            if td.get_text().find("직위명")>=0:
                chair=tds[index+1].get_text()
                print("chair:",chair,"/ chair_TYPE:",type(chair))
        
        data={'name':name,'chair':chair}
        return data
    elif data['text'].find("특정증권등의") >= 0:
        tds = soup.find_all("td")
        totalValue = 0
        regiDateList=[]
        totalAmount=0

        buyCount=0
        sellCount=0
        for index,td in enumerate(tds):
            # print("td.get_text():",td.get_text(),"/ td.get_text()_TYPE:",type(td.get_text()))
            if td.get_text().find("장내매수")>=0 and td.get_text().find("이상")<0:
                print('4124124')
                try:
                    # changeAmt = int(tds[index + 4].get_text().strip().replace(",", ""))
                    changeAmt = int(re.sub(r'[^0-9]', '', tds[index + 4].get_text().strip()))
                    print("changeAmt:", changeAmt, "/ changeAmt_TYPE:", type(changeAmt))
                except:
                    changeAmt=0
                print("changeAmt:",changeAmt)
                if changeAmt==0:
                    continue
                try:
                    # changePrice = int(tds[index + 6].get_text().strip().replace(",", ""))
                    changePrice = int(re.sub(r'[^0-9]', '', tds[index + 6].get_text().strip()))
                    print("changePrice:", changePrice, "/ changePrice_TYPE:", type(changePrice))

                except:
                    changePrice=0
                print("changePrice:",changePrice)

                changeValue=changeAmt*changePrice
                print("changeValue:",changeValue,"/ changeValue_TYPE:",type(changeValue))
                totalValue=totalValue+changeValue
                totalAmount=totalAmount+changeAmt
                regiDate=tds[index+1].get_text()
                print("regiDate:",regiDate,"/ regiDate_TYPE:",type(regiDate))
                regiDateList.append(regiDate)
                buyCount+=1
                print("======================================")
            elif td.get_text().find("장내매도")>=0 and td.get_text().find("이상")<0:
                print('1234')
                try:
                    # changeAmt=int(tds[index+4].get_text().strip().replace(",",""))
                    # 정규표현식을 사용하여 쉼표 및 숫자 이외의 문자 제거
                    changeAmt=int(re.sub(r'[^0-9]', '', tds[index+4].get_text().strip()))
                except:
                    continue
                print("changeAmt:",changeAmt,"/ changeAmt_TYPE:",type(changeAmt))
                try:
                    # changePrice=int(tds[index+6].get_text().strip().replace(",","").replace("처분","").replace("(","").replace(")","").replace("원",""))
                    print('changePrice11',tds[index+6].get_text().strip())
                    changePrice=int(re.sub(r'[^0-9]', '', tds[index+6].get_text().strip()))
                except:
                    print("에러인듯")
                    continue
                print("changePrice:",changePrice,"/ changePrice_TYPE:",type(changePrice))
                changeValue=changeAmt*changePrice
                print("changeValue:",changeValue,"/ changeValue_TYPE:",type(changeValue))
                totalValue=totalValue+changeValue
                totalAmount=totalAmount+changeAmt
                regiDate=tds[index+1].get_text()
                print("regiDate:",regiDate,"/ regiDate_TYPE:",type(regiDate))
                regiDateList.append(regiDate)
                sellCount+=1
                print("======================================")
            elif td.get_text().find("이번보고서")>=0:
                print("5124312")
                try:
                    thisStock=tds[index+4].get_text().strip()
                except:
                    thisStock=""
                print("thisStock:",thisStock)
                try:
                    thisRatio=tds[index+5].get_text().strip()
                except:
                    thisRatio=""
                print("thisRatio:",thisRatio)
                print("======================================")
            elif td.get_text().find("증     감")>=0:
                print('429304')
                try:
                    JGStock=tds[index+3].get_text().strip()
                except:
                    JGStock=""
                print("JGStock:",JGStock)
                try:
                    JGRatio=tds[index+4].get_text().strip()
                except:
                    JGRatio=""
                print("JGRatio:",JGRatio)
                print("======================================")

        tables=soup.find_all("table")
        totalStock=''
        for index,table in enumerate(tables):
            try:
                if table.find('th').get_text().find("발행주식 총수")>=0:
                    totalStock=table.find('td').get_text()
                    print("totalStock:",totalStock)
                else:
                    print('없음')
            except:
                totalStock = ""



        if buyCount>=1 and sellCount>=1:
            saleType="장내매수매도"
        elif buyCount>=1 and sellCount==0:
            saleType="장내매수"
        elif sellCount>=1 and buyCount==0:
            saleType="장내매도"
        else:
            saleType="없음"


        totalValue="{:,}".format(totalValue)+"원"
        print('123')
        thisRatio=thisRatio+"%"
        print('23589')
        JGRatio=JGRatio+"%"
        print("234")

        print("totalValue:",totalValue,"/ totalValue_TYPE:",type(totalValue))
        print("regiDateList:",regiDateList,"/ regiDateList_TYPE:",type(regiDateList))
        print("saleType:",saleType,"/ saleType_TYPE:",type(saleType))

        data={'totalValue':totalValue,'regiDateList':regiDateList,'totalAmount':totalAmount,'saleType':saleType,'totalStock':totalStock,'thisStock':thisStock,'thisRatio':thisRatio,'JGStock':JGStock,'JGRatio':JGRatio}
        pprint.pprint(data)
        return data
    else:
        data={}
        return data

def GetInfos2(idInfo):
    cookies = {
        'WMONID': 'D1vNDyvY7sc',
        'JSESSIONID': 'ZdcjGEP9NXaljIkvkDg1eCpLRO65fgd5fZb1JMklaT1wHkBPn1XXsHBxs9FEzFH1.ZG1fZGFydC9kYXJ0MV9kYXJ0X21zMw==',
    }

    headers = {
        'Accept': 'text/html,application/xhtml+xml,application/xml;q=0.9,image/avif,image/webp,image/apng,*/*;q=0.8,application/signed-exchange;v=b3;q=0.7',
        'Accept-Language': 'ko-KR,ko;q=0.9,en-US;q=0.8,en;q=0.7',
        'Cache-Control': 'max-age=0',
        'Connection': 'keep-alive',
        # 'Cookie': 'WMONID=D1vNDyvY7sc; JSESSIONID=ZdcjGEP9NXaljIkvkDg1eCpLRO65fgd5fZb1JMklaT1wHkBPn1XXsHBxs9FEzFH1.ZG1fZGFydC9kYXJ0MV9kYXJ0X21zMw==',
        'Referer': 'https://dart.fss.or.kr/dsab007/main.do',
        'Sec-Fetch-Dest': 'document',
        'Sec-Fetch-Mode': 'navigate',
        'Sec-Fetch-Site': 'same-origin',
        'Sec-Fetch-User': '?1',
        'Upgrade-Insecure-Requests': '1',
        'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/120.0.0.0 Safari/537.36',
        'sec-ch-ua': '"Not_A Brand";v="8", "Chromium";v="120", "Google Chrome";v="120"',
        'sec-ch-ua-mobile': '?0',
        'sec-ch-ua-platform': '"Windows"',
    }

    params = {
        'rcpNo': idInfo['reportId'],
    }
    print("reportId:",idInfo['reportId'])

    response = requests.get('https://dart.fss.or.kr/dsaf001/main.do', params=params, cookies=cookies, headers=headers)
    soup=BeautifulSoup(response.text,'lxml')
    btnDown=soup.find("button",attrs={'class':'btnDown'})['onclick']
    regex=re.compile('\d+')
    dcmNo=regex.findall(btnDown)[-1]
    print("dcmNo:",dcmNo,"/ dcmNo_TYPE:",type(dcmNo))

    cookies = {
        'WMONID': 'D1vNDyvY7sc',
        'JSESSIONID': 'ZdcjGEP9NXaljIkvkDg1eCpLRO65fgd5fZb1JMklaT1wHkBPn1XXsHBxs9FEzFH1.ZG1fZGFydC9kYXJ0MV9kYXJ0X21zMw==',
    }

    headers = {
        'Accept': 'text/html,application/xhtml+xml,application/xml;q=0.9,image/avif,image/webp,image/apng,*/*;q=0.8,application/signed-exchange;v=b3;q=0.7',
        'Accept-Language': 'ko-KR,ko;q=0.9,en-US;q=0.8,en;q=0.7',
        'Connection': 'keep-alive',
        # 'Cookie': 'WMONID=D1vNDyvY7sc; JSESSIONID=ZdcjGEP9NXaljIkvkDg1eCpLRO65fgd5fZb1JMklaT1wHkBPn1XXsHBxs9FEzFH1.ZG1fZGFydC9kYXJ0MV9kYXJ0X21zMw==',
        'Referer': 'https://dart.fss.or.kr/dsaf001/main.do?rcpNo=20231214801335',
        'Sec-Fetch-Dest': 'iframe',
        'Sec-Fetch-Mode': 'navigate',
        'Sec-Fetch-Site': 'same-origin',
        'Upgrade-Insecure-Requests': '1',
        'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/120.0.0.0 Safari/537.36',
        'sec-ch-ua': '"Not_A Brand";v="8", "Chromium";v="120", "Google Chrome";v="120"',
        'sec-ch-ua-mobile': '?0',
        'sec-ch-ua-platform': '"Windows"',
    }

    params = {
        'rcpNo': idInfo['reportId'],
        'dcmNo': dcmNo,
        'eleId': '0',
        'offset': '0',
        'length': '0',
        'dtd': 'HTML',
    }

    response = requests.get('https://dart.fss.or.kr/report/viewer.do', params=params, cookies=cookies, headers=headers)
    soup=BeautifulSoup(response.text,'lxml')

    totalResult=[]

    firstTable=soup.find("table",attrs={'id':'XFormD1_Form0_RepeatTable0'})
    firstTableAdds = soup.find_all("table", attrs={'id': 'XFormG1_Form0_RepeatTable0'})
    if len(firstTableAdds) >= 1:
        firstTables=[firstTable]+firstTableAdds
    else:
        firstTables=[firstTable]
    for firstTable in firstTables:


        allTds=firstTable.find_all('td')
        for index,td in enumerate(allTds):
            if td.get_text().find("성명")>=0:
                company=allTds[index+1].get_text().strip()
                print("company:",company,"/ company_TYPE:",type(company))
            if td.get_text().find("발행회사와의 관계") >= 0:
                relation=allTds[index+1].get_text().strip()
                print("relation:",relation,"/ relation_TYPE:",type(relation))
        companyRelation=company+" / "+relation

        thirdTable=soup.find("table",attrs={'id':'XFormD1_Form0_Table1'})
        totalStock=thirdTable.find_all('td')[-1].get_text().strip()
        print("totalStock:",totalStock,"/ totalStock_TYPE:",type(totalStock))


        changeReasonList=[]
        changeReasonValueList=[]
        firstTableTrs=firstTable.find_all('tr')
        for index,firstTableTr in enumerate(firstTableTrs):
            if index>=4:
                changeReason=firstTableTr.find_all('td')[1].get_text().strip()
                changeReasonValue = int(firstTableTr.find_all('td')[4].get_text().strip().replace(",",""))
                changeReasonList.append(changeReason)
                changeReasonValueList.append(changeReasonValue)



        changeReason=list(set(changeReasonList))
        changeReason=",".join(changeReason)
        print("changeReason:",changeReason,"/ changeReason_TYPE:",type(changeReason))

        changeReasonValueTotal=0
        # 리스트의 모든 요소를 더하기
        for number in changeReasonValueList:
            changeReasonValueTotal += number
        print("changeReasonValueTotal:",changeReasonValueTotal,"/ changeReasonValueTotal_TYPE:",type(changeReasonValueTotal))

        changeReasonValueTotalRatio=format_number(changeReasonValueTotal/int(totalStock.replace(",",""))*100)

        print("changeReasonValueTotalRatio:",changeReasonValueTotalRatio,"/ changeReasonValueTotalRatio_TYPE:",type(changeReasonValueTotalRatio))
        # 숫자를 세 자리 쉼표로 처리하는 문자열 생성
        changeReasonValueTotal = "{:,}".format(changeReasonValueTotal)

        sumList=[]
        ratioList=[]
        secondTable=soup.find("table",attrs={'id':'XFormD1_Form0_Table2'})
        secondTableTds=secondTable.find_all('td')
        for index,secondTableTd in enumerate(secondTableTds):
            if secondTableTd.get_text().find("합계")>=0:
                sumValue=secondTableTds[index+1].get_text()
                ratioValue = secondTableTds[index + 2].get_text()
                sumList.append(sumValue)
                ratioList.append(ratioValue)

        try:
            thisSum=sumList[1].strip()

        except:
            thisSum=""
        print("thisSum:",thisSum)
        try:
            changeSum=sumList[2].strip()
        except:
            changeSum=""
        print("changeSum:",changeSum)

        try:
            thisRatio=ratioList[1].strip()
        except:
            thisRatio=""
        print("thisRatio:",thisRatio)

        try:
            changeRatio=ratioList[2].strip()
        except:
            changeRatio=""
        print("changeRatio:",changeRatio)



        fourthTable=soup.find("table",attrs={'id':'XFormD1_Form0_RepeatTable1'})
        allTrs=fourthTable.find_all('tr')
        companyStock=""
        companyRatio=""
        for tr in allTrs:
            # print(tr.find_all('td')[0].get_text())
            if tr.find_all('td')[0].get_text().find(company.replace(" ",""))>=0:
                companyStock=tr.find_all('td')[-2].get_text().strip()
                print("companyStock:",companyStock,"/ companyStock_TYPE:",type(companyStock))
                companyRatio=tr.find_all('td')[-1].get_text().strip()
                print("companyRatio:",companyRatio,"/ companyRatio_TYPE:",type(companyRatio))

        url='https://dart.fss.or.kr/dsaf001/main.do?rcpNo={}'.format(idInfo['reportId'])
        result={'companyRelation':companyRelation,'changeReason':changeReason,'totalStock':totalStock+" 주","stock1":changeReasonValueTotal+" 주\n"+"("+changeReasonValueTotalRatio+"%)","stock2":thisSum+" 주\n"+"("+thisRatio+"%)","stock3":companyStock+" 주\n"+"("+companyRatio+"%)","url":url}
        pprint.pprint(result)
        totalResult.append(result)
        print("=====================")
    return totalResult
    # return companyRelation,changeReason,totalStock,changeSum+"\n"+"("+changeRatio+"%)",thisSum+"\n"+"("+thisRatio+"%)",companyStock+"\n"+"("+companyRatio+"%)",url

# idInfo={
#     "title": "한국앤컴퍼니",
#     "reportId": "20231215800001",
#     "regiDate": "2023.12.14"
#   }
# GetInfos2(idInfo)

class Thread(QThread):
    cnt = 0
    user_signal = pyqtSignal(str)  # 사용자 정의 시그널 2 생성

    def __init__(self, parent,year,month,day,category):  # parent는 WndowClass에서 전달하는 self이다.(WidnowClass의 인스턴스)
        super().__init__(parent)
        self.parent = parent  # self.parent를 사용하여 WindowClass 위젯을 제어할 수 있다.
        self.year=year
        self.month=month
        self.day=day
        self.category=category
    def run(self):
        text = "ID가져오기 시작"
        self.user_signal.emit(text)
        # 엑셀양식만들기

        wb = openpyxl.Workbook()
        ws = wb.active
        if self.category=="임원":
            title = ['날짜','기업명','매매임원','매수/매도','이번보고서 주식수(주)','변동주식 수(주)','발행주식 총수(주)','비율(E/F)','주식비율(D/F)','총금액', '변동일', '링크']
        else:
            title = ['날짜', '기업명', '보고자 성명 / 최대주주 및 발행회사와의 관계','변경원인','발행주식총수','증감 주식수','증감 후 주식수 (보고자만)','증감 후 주식수 (전체)','링크']
        ws.append(title)

        # ================id가져오기
        idList = GetID(self.category,self.year, self.month, self.day)
        text = "ID가져오기 완료"
        self.user_signal.emit(text)
        # ================
        with open('idList.json', "r", encoding='utf-8-sig') as f:
            idList = json.load(f)

        # idList=idList[:3]

        timeNow = datetime.datetime.now().strftime("%Y%m%d_%H%M%S")
        # =============상세정보가져오기
        for index,idInfo in enumerate(idList):
            text = "{}/{}({})번째 확인중...".format(index+1,len(idList),idInfo['title'])
            self.user_signal.emit(text)
            #=============임원
            if self.category=="임원":
                dataList = GetPreDetali(idInfo)
                with open('dataList.json', 'w', encoding='utf-8-sig') as f:
                    json.dump(dataList, f, indent=2, ensure_ascii=False)
                results = {}
                results.update({'reportDate': idInfo['regiDate']})
                for data in dataList:
                    try:
                        result = GetInfos(data)
                        time.sleep(random.randint(5,10)*0.1)
                    except:
                        print("에러22")
                    results.update(result)
                    time.sleep(0.5)
                print("results:",results,"/ results_TYPE:",type(results))
                if len(results['regiDateList']) >= 2:
                    period = results['regiDateList'][0] + "~" + results['regiDateList'][-1]
                    text = "거래있음"
                    self.user_signal.emit(text)
                    dataList = GetPreDetali(idInfo)
                elif len(results['regiDateList']) == 1:
                    period = results['regiDateList'][0]
                    text = "거래있음"
                    self.user_signal.emit(text)
                else:
                    text = "거래없음"
                    self.user_signal.emit(text)
                    print(text)
                    continue
                print("데이타저장")
                pprint.pprint(results)
                dataRow = [results['reportDate'],idInfo['title'],results['chair']+","+results['name'],results['saleType'], results['thisStock'],results['JGStock'],results['totalStock'],results['JGRatio'],
                           results['thisRatio'],results['totalValue'],period,results['url']]
                print("dataRow:",dataRow,"/ dataRow_TYPE:",type(dataRow))
                ws.append(dataRow)
                wb.save('전자공시조회_{}_{}.xlsx'.format(self.category,timeNow))
            #==============최대주주
            else:
                # dataList = GetPreDetali(idInfo)
                # with open('dataList.json', 'w', encoding='utf-8-sig') as f:
                #     json.dump(dataList, f, indent=2, ensure_ascii=False)
                results = {}
                results.update({'reportDate': idInfo['regiDate']})
                try:
                    resultInfos=GetInfos2(idInfo)
                except:
                    print("에러발생")
                    continue
                time.sleep(random.randint(5, 10) * 0.1)

                for resultInfo in resultInfos:
                    results.update(resultInfo)
                    # for data in dataList:
                    #     try:
                    #         result = GetInfos(data)
                    #         time.sleep(random.randint(5, 10) * 0.1)
                    #     except:
                    #         print("에러22")
                    #     results.update(result)
                    #     time.sleep(0.5)

                    dataRow = [results['reportDate'], idInfo['title'], results['companyRelation'],
                               results['changeReason'], results['totalStock'], results['stock1'],results['stock3'], results['stock2'],
                               results['url']]
                    print("dataRow:", dataRow, "/ dataRow_TYPE:", type(dataRow))
                    ws.append(dataRow)
                    wb.save('전자공시조회_{}_{}.xlsx'.format(self.category,timeNow))

                # 두 번째 열('B')을 검사하고 연속되는 셀 합치기
                previous_value = None
                start_row = None

                for row in range(1, ws.max_row + 1):
                    current_value = ws[f'B{row}'].value
                    if current_value == previous_value and current_value is not None:
                        if start_row is None:
                            start_row = row - 1
                    else:
                        if start_row is not None:
                            ws.merge_cells(start_row=start_row, start_column=2, end_row=row - 1, end_column=2)
                            start_row = None
                    previous_value = current_value

                # 마지막 연속된 셀 그룹을 확인 및 합치기
                if start_row is not None:
                    ws.merge_cells(start_row=start_row, start_column=2, end_row=ws.max_row, end_column=2)



        # 전체 행에 자동 줄바꿈 설정
        for index,row in enumerate(ws.iter_rows()):
            if index==0:
                for index,cell in enumerate(row):
                    cell.alignment = Alignment(wrapText=True,vertical='center',horizontal='center')
            else:
                for index,cell in enumerate(row):
                    # if index<=4:
                    cell.alignment = Alignment(wrapText=True,vertical='center',horizontal='center')
                    # elif 5<=index<=6:
                    #     cell.alignment = Alignment(wrapText=True,vertical='top')
                    # else:
                    #     cell.alignment = Alignment(wrapText=True, vertical='center', horizontal='center')

        # 행 높이 지정
        for i in range(2, ws.max_row + 1):
            ws.row_dimensions[i].height = 30
        if self.category=="임원":
            # 열의 폭 설정
            ws.column_dimensions['A'].width = 15
            ws.column_dimensions['B'].width = 15
            ws.column_dimensions['C'].width = 15
            ws.column_dimensions['D'].width = 15
            ws.column_dimensions['E'].width = 15
            ws.column_dimensions['F'].width = 15
            ws.column_dimensions['G'].width = 15
            ws.column_dimensions['H'].width = 15
            ws.column_dimensions['I'].width = 15
            ws.column_dimensions['J'].width = 40
            ws.column_dimensions['K'].width = 15
            ws.column_dimensions['L'].width = 15

            # 데이터 읽기 (헤더 제외)
            rows = list(ws.iter_rows(values_only=True))
            header, data = rows[0], rows[1:]

            print("data:", data, "/ data_TYPE:", type(data))

            # '장내매수', '장내매수매도', '장내매도' 행을 분리하여 정렬
            buyers = [row for row in data if row[3] == '장내매수']
            buyers_sellers = [row for row in data if row[3] == '장내매수매도']
            sellers = [row for row in data if row[3] == '장내매도']

            # 데이터를 다시 합치기
            sorted_data = buyers + buyers_sellers + sellers

            # 기존 데이터 지우기
            for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
                for cell in row:
                    cell.value = None
            #
            # 정렬된 데이터 쓰기
            for row_index, row in enumerate(sorted_data, start=2):
                for col_index, value in enumerate(row, start=1):
                    ws.cell(row=row_index, column=col_index, value=value)

            # 각 조건에 따른 글자색 설정
            colors = {
                '장내매수': 'FF0000',  # 빨강
                '장내매도': '0000FF',  # 파랑
                '장내매수매도': '00FF00'  # 초록
            }

            # 3번째 열의 셀을 순회하며 글자색 변경
            for row in ws.iter_rows(min_row=2, max_col=4, max_row=ws.max_row):
                cell = row[3]  # 3번째 열의 셀
                if cell.value in colors:
                    cell.font = Font(color=colors[cell.value])

        else:
            ws.column_dimensions['A'].width = 20
            ws.column_dimensions['B'].width = 20
            ws.column_dimensions['C'].width = 40
            ws.column_dimensions['D'].width = 20
            ws.column_dimensions['E'].width = 20
            ws.column_dimensions['F'].width = 20
            ws.column_dimensions['G'].width = 20
            ws.column_dimensions['H'].width = 20
            ws.column_dimensions['I'].width = 20

        # 첫 번째 행에 회색 배경 색상 적용
        for cell in ws[1]:  # 첫 번째 행의 모든 셀에 스타일 적용
            cell.fill = PatternFill(start_color="A9A9A9", end_color="A9A9A9", fill_type="solid")  # 회색으로 설정

        # 첫 번째 행 고정
        ws.freeze_panes = 'A1'

        # # 전체 열에 필터 적용
        ws.auto_filter.ref = ws.dimensions

        wb.save('전자공시조회_{}_{}.xlsx'.format(self.category,timeNow))

        text = "작업완료"
        self.user_signal.emit(text)

    def stop(self):
        pass


class Example(QMainWindow, Ui_MainWindow):
    def __init__(self):
        super().__init__()
        self.path = "C:"
        self.index = None
        self.setupUi(self)
        self.setSlot()
        self.show()
        QApplication.processEvents()
        today=QDate.currentDate()
        self.dateEdit.setDate(today)

    def start(self):
        self.category=self.comboBox.currentText()
        self.year=self.dateEdit.date().year()
        self.month = self.dateEdit.date().month()
        self.day = self.dateEdit.date().day()
        self.x = Thread(self,self.year,self.month,self.day,self.category)
        self.x.user_signal.connect(self.slot1)  # 사용자 정의 시그널2 슬롯 Connect
        self.x.start()

    def slot1(self, data1):  # 사용자 정의 시그널1에 connect된 function
        self.textEdit.append(str(data1))

    def setSlot(self):
        pass

    def setIndex(self, index):
        pass

    def quit(self):
        QCoreApplication.instance().quit()


app = QApplication([])
ex = Example()
sys.exit(app.exec_())